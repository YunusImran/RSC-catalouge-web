"""Import catalog data from /tmp/catalog.xlsx, replacing existing data.

Strategy:
- Read each row from row 3 onward of the Excel.
- Use 'Sr. No' as the unique catalog code (RSC-NNNNN).
- Use CATALOUGE NAME as catalog_name (fallback: '—' or CAT NO).
- Auto-create suppliers by name.
- Wipe existing catalogs, catalog_issues, catalog_returns, scan_history first (clean import).
"""
import asyncio
import os
from datetime import datetime, timezone
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook

load_dotenv(Path(__file__).parent / ".env")
from motor.motor_asyncio import AsyncIOMotorClient


def s(v):
    return str(v).strip() if v is not None else ""


async def main():
    client = AsyncIOMotorClient(os.environ["MONGO_URL"])
    db = client[os.environ["DB_NAME"]]

    # 1) Wipe existing catalog data + related transactions
    print("Wiping existing catalogs/issues/returns/scans...")
    for col in ["catalogs", "catalog_issues", "catalog_returns", "scan_history", "import_logs", "counters"]:
        r = await db[col].delete_many({})
        print(f"  {col}: deleted {r.deleted_count}")

    # 2) Load suppliers cache
    suppliers = {}
    async for s_doc in db.suppliers.find({}):
        suppliers[s_doc["name"].strip().lower()] = str(s_doc["_id"])

    # 3) Parse Excel
    wb = load_workbook("/tmp/catalog.xlsx", data_only=True)
    ws = wb.active
    now_iso = datetime.now(timezone.utc).isoformat()

    inserted = 0
    skipped = 0
    new_suppliers = 0
    docs = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        sr_no, name, cat_no, qty, supp, remarks, recv_date = row[:7]
        # Skip totally empty rows
        if not sr_no and not name and not supp and not cat_no:
            continue
        if sr_no is None:
            skipped += 1
            continue

        try:
            sr_int = int(sr_no)
        except (TypeError, ValueError):
            skipped += 1
            continue

        code = f"RSC-{sr_int:05d}"
        catalog_name = s(name) or s(cat_no) or "—"
        cat_no_val = s(cat_no)
        supplier_name = s(supp)
        supplier_id = None
        if supplier_name:
            key = supplier_name.lower()
            if key not in suppliers:
                res = await db.suppliers.insert_one({
                    "name": supplier_name, "is_archived": False,
                    "created_at": now_iso, "updated_at": now_iso, "seeded": True,
                })
                suppliers[key] = str(res.inserted_id)
                new_suppliers += 1
            supplier_id = suppliers[key]

        # parse date
        recv_iso = ""
        if recv_date is not None:
            if isinstance(recv_date, datetime):
                recv_iso = recv_date.date().isoformat()
            else:
                recv_iso = s(recv_date)

        try:
            quantity = int(qty) if qty not in (None, "") else 1
        except (TypeError, ValueError):
            quantity = 1

        docs.append({
            "catalog_code": code,
            "catalog_name": catalog_name,
            "cat_no": cat_no_val,
            "quantity": quantity,
            "supplier_id": supplier_id,
            "category_id": None,
            "fabric_type": "", "material_composition": "",
            "gsm": None, "color": "",
            "total_swatches": 0,
            "description": "",
            "catalog_image": "",
            "swatch_images": [],
            "qr_value": code,
            "buying_price": None, "selling_price": None,
            "remarks": s(remarks),
            "receiving_date": recv_iso,
            "barcode_value": code,
            "status": "Available",
            "is_archived": False,
            "created_at": now_iso, "updated_at": now_iso,
            "created_by": "import",
            "imported": True,
            "sr_no": sr_int,
        })

    # 4) Bulk insert
    if docs:
        BATCH = 500
        for i in range(0, len(docs), BATCH):
            chunk = docs[i:i + BATCH]
            await db.catalogs.insert_many(chunk)
            inserted += len(chunk)
            print(f"  inserted {inserted}/{len(docs)}")

    # 5) Reset transaction counter
    await db.counters.update_one({"_id": "transactions"}, {"$set": {"value": 0}}, upsert=True)

    print(f"\nDone. Inserted {inserted} catalogs · {new_suppliers} new suppliers · {skipped} skipped rows.")
    client.close()


asyncio.run(main())
