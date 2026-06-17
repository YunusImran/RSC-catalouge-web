"""Import catalog data from /tmp/catalog_v5.xlsx (v5 file with pre-made barcodes in col B).

Columns (row 2 header):
  A: Sr. No
  B: RSC-NNNNN  (pre-made barcode/code)
  C: CATALOUGE NAME
  D: CAT NO
  E: QUANTITY
  F: SUPPLIER NAME
  G: REMARKS
  H: RECEVING DATE

Strategy: column B value is BOTH the catalog_code AND the barcode_value AND the qr_value.
"""
import asyncio
import os
from datetime import datetime, timezone
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook

load_dotenv(Path(__file__).parent / ".env")
from motor.motor_asyncio import AsyncIOMotorClient


def s(v):
    return str(v).strip() if v is not None else ""


async def main():
    client = AsyncIOMotorClient(os.environ["MONGO_URL"])
    db = client[os.environ["DB_NAME"]]

    print("Wiping catalogs/issues/returns/scans/import_logs/counters...")
    for col in ["catalogs", "catalog_issues", "catalog_returns", "scan_history", "import_logs", "counters"]:
        r = await db[col].delete_many({})
        print(f"  {col}: deleted {r.deleted_count}")

    # supplier cache (keep existing suppliers)
    suppliers = {}
    async for s_doc in db.suppliers.find({}):
        suppliers[s_doc["name"].strip().lower()] = str(s_doc["_id"])

    wb = load_workbook("/tmp/catalog_v5.xlsx", data_only=True)
    ws = wb.active
    now_iso = datetime.now(timezone.utc).isoformat()

    inserted = 0
    skipped = 0
    new_suppliers = 0
    seen_codes = set()
    docs = []

    # Data starts at row 3 (row 1 = "CATALOUGE" title, row 2 = headers)
    for row in ws.iter_rows(min_row=3, values_only=True):
        sr_no, code_val, name, cat_no, qty, supp, remarks, recv_date = row[:8]
        if not code_val and not name and not supp and not cat_no:
            continue

        code = s(code_val)
        if not code:
            skipped += 1
            continue
        if code in seen_codes:
            skipped += 1
            continue
        seen_codes.add(code)

        catalog_name = s(name) or s(cat_no) or "—"
        cat_no_val = s(cat_no)
        supplier_name = s(supp)
        supplier_id = None
        if supplier_name:
            key = supplier_name.lower()
            if key not in suppliers:
                res = await db.suppliers.insert_one({
                    "name": supplier_name, "is_archived": False,
                    "created_at": now_iso, "updated_at": now_iso, "seeded": True,
                })
                suppliers[key] = str(res.inserted_id)
                new_suppliers += 1
            supplier_id = suppliers[key]

        recv_iso = ""
        if recv_date is not None:
            if isinstance(recv_date, datetime):
                recv_iso = recv_date.date().isoformat()
            else:
                recv_iso = s(recv_date)

        try:
            quantity = int(qty) if qty not in (None, "") else 1
        except (TypeError, ValueError):
            quantity = 1

        try:
            sr_int = int(sr_no) if sr_no is not None else 0
        except (TypeError, ValueError):
            sr_int = 0

        docs.append({
            "catalog_code": code,        # from col B
            "barcode_value": code,        # SAME barcode from the file
            "qr_value": code,             # also the QR scannable value
            "catalog_name": catalog_name,
            "cat_no": cat_no_val,
            "quantity": quantity,
            "supplier_id": supplier_id,
            "category_id": None,
            "fabric_type": "", "material_composition": "",
            "gsm": None, "color": "",
            "total_swatches": 0,
            "description": "",
            "catalog_image": "",
            "swatch_images": [],
            "buying_price": None, "selling_price": None,
            "remarks": s(remarks),
            "receiving_date": recv_iso,
            "status": "Available",
            "is_archived": False,
            "created_at": now_iso, "updated_at": now_iso,
            "created_by": "import",
            "imported": True,
            "sr_no": sr_int,
        })

    if docs:
        BATCH = 500
        for i in range(0, len(docs), BATCH):
            chunk = docs[i:i + BATCH]
            await db.catalogs.insert_many(chunk)
            inserted += len(chunk)
            print(f"  inserted {inserted}/{len(docs)}")

    await db.counters.update_one({"_id": "transactions"}, {"$set": {"value": 0}}, upsert=True)

    print(f"\nDone. Inserted {inserted} catalogs · {new_suppliers} new suppliers · {skipped} skipped rows.")
    client.close()


asyncio.run(main())
