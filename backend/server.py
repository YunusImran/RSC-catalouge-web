from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import csv
import secrets
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, Literal

import bcrypt
import jwt
import barcode as pybarcode
from barcode.writer import SVGWriter
import qrcode
import qrcode.image.svg
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel, EmailStr
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId


# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
JWT_ALGORITHM = "HS256"
ACCESS_TTL_MIN = 60 * 8
REFRESH_TTL_DAYS = 7
ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_STAFF = "admin", "supervisor", "staff"
ALL_ROLES = (ROLE_ADMIN, ROLE_SUPERVISOR, ROLE_STAFF)
MOBILE_RE = re.compile(r"^[+]?[0-9\-\s()]{7,20}$")

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Royal Shades Catalog API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(name)s - %(message)s')
log = logging.getLogger("rsc")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def now() -> datetime:
    return datetime.now(timezone.utc)

def iso(dt: datetime) -> str:
    return dt.isoformat()

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def create_access_token(user_id: str, email: str, role: str) -> str:
    return jwt.encode(
        {"sub": user_id, "email": email, "role": role,
         "exp": now() + timedelta(minutes=ACCESS_TTL_MIN), "type": "access"},
        get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    return jwt.encode({"sub": user_id, "exp": now() + timedelta(days=REFRESH_TTL_DAYS), "type": "refresh"},
                      get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax",
                        max_age=ACCESS_TTL_MIN * 60, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=False, samesite="lax",
                        max_age=REFRESH_TTL_DAYS * 86400, path="/")

def clear_auth_cookies(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")

def doc_to_json(doc):
    if not doc:
        return doc
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc.pop("_id"))
    for k, v in list(doc.items()):
        if isinstance(v, ObjectId):
            doc[k] = str(v)
        if isinstance(v, datetime):
            doc[k] = v.isoformat()
    return doc

def strip_buying_price(catalog: dict, role: str) -> dict:
    if role != ROLE_ADMIN and catalog:
        catalog = dict(catalog)
        catalog.pop("buying_price", None)
    return catalog

async def next_transaction_id() -> str:
    """Returns next sequential transaction ID like TXN-0001 (shared issues+returns)."""
    from pymongo import ReturnDocument
    res = await db.counters.find_one_and_update(
        {"_id": "transactions"},
        {"$inc": {"value": 1}},
        upsert=True, return_document=ReturnDocument.AFTER
    )
    n = (res or {}).get("value", 1)
    return f"TXN-{int(n):04d}"

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if not user.get("is_active", True):
            raise HTTPException(status_code=403, detail="Account disabled")
        user["id"] = str(user.pop("_id"))
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles: str):
    async def checker(user: dict = Depends(get_current_user)):
        if user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return checker

async def audit(user, action, description, request=None, affected=""):
    ip = (request.headers.get("x-forwarded-for", "").split(",")[0].strip()
          if request else "n/a") or (request.client.host if request and request.client else "n/a")
    await db.audit_logs.insert_one({
        "user_id": user["id"], "user_email": user["email"], "user_name": user.get("name", ""),
        "action": action, "description": description, "record_affected": affected,
        "ip_address": ip, "created_at": iso(now())
    })


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class LoginIn(BaseModel):
    username: str
    password: str

class RegisterIn(BaseModel):
    username: str
    email: Optional[str] = ""
    password: str
    name: str
    role: Literal["admin", "supervisor", "staff"] = "staff"

class ChangePasswordIn(BaseModel):
    old_password: str
    new_password: str

class ForgotPasswordIn(BaseModel):
    email: EmailStr

class ResetPasswordIn(BaseModel):
    token: str
    new_password: str

class CategoryIn(BaseModel):
    name: str
    description: Optional[str] = ""

class SupplierIn(BaseModel):
    name: str
    contact_person: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    gst_number: Optional[str] = ""
    notes: Optional[str] = ""

class EmployeeIn(BaseModel):
    name: str
    employee_code: Optional[str] = ""
    department: Optional[str] = ""
    designation: Optional[str] = ""
    mobile: Optional[str] = ""
    email: Optional[str] = ""
    is_active: bool = True

class CatalogIn(BaseModel):
    catalog_code: str
    catalog_name: str
    cat_no: Optional[str] = ""
    quantity: Optional[int] = 1
    receiving_date: Optional[str] = ""
    remarks: Optional[str] = ""
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    fabric_type: Optional[str] = ""
    material_composition: Optional[str] = ""
    gsm: Optional[float] = None
    color: Optional[str] = ""
    total_swatches: Optional[int] = 0
    description: Optional[str] = ""
    catalog_image: Optional[str] = ""
    qr_value: Optional[str] = ""
    buying_price: Optional[float] = None
    selling_price: Optional[float] = None
    swatch_images: Optional[list] = None

class IssueIn(BaseModel):
    catalog_id: str
    customer_name: Optional[str] = ""
    employee_id: Optional[str] = None
    employee_name: Optional[str] = ""
    department: Optional[str] = ""
    mobile: str       # MANDATORY (validated below)
    email: Optional[str] = ""
    issue_date: Optional[str] = None
    expected_return_date: Optional[str] = None
    remarks: Optional[str] = ""

class ReturnIn(BaseModel):
    catalog_id: str
    returned_by: Optional[str] = ""
    return_date: Optional[str] = None
    condition: Literal["Excellent", "Good", "Damaged", "Missing Swatches"] = "Good"
    remarks: Optional[str] = ""

class ScanIn(BaseModel):
    barcode_value: str
    device_type: Optional[str] = "Web"
    action: Literal["Search", "Issue", "Return", "View"] = "Search"
    remarks: Optional[str] = ""


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
@api.post("/auth/login")
async def login(body: LoginIn, response: Response, request: Request):
    username = body.username.lower().strip()
    identifier = f"user:{username}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("locked_until"):
        locked_until = datetime.fromisoformat(attempt["locked_until"])
        if locked_until.tzinfo is None:
            locked_until = locked_until.replace(tzinfo=timezone.utc)
        if locked_until > now():
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again later.")

    # match by username OR email (back-compat)
    user = await db.users.find_one({"$or": [{"username": username}, {"email": username}]})
    if not user or not verify_password(body.password, user["password_hash"]):
        attempts = (attempt.get("count", 0) if attempt else 0) + 1
        update = {"count": attempts, "updated_at": iso(now())}
        if attempts >= 5:
            update["locked_until"] = iso(now() + timedelta(minutes=15))
            update["count"] = 0
        await db.login_attempts.update_one({"identifier": identifier},
                                           {"$set": {"identifier": identifier, **update}}, upsert=True)
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account disabled")
    await db.login_attempts.delete_one({"identifier": identifier})

    uid = str(user["_id"])
    access = create_access_token(uid, user.get("email") or username, user["role"])
    refresh = create_refresh_token(uid)
    set_auth_cookies(response, access, refresh)
    user["id"] = uid
    user.pop("_id"); user.pop("password_hash", None)
    await db.audit_logs.insert_one({"user_id": uid, "user_email": user.get("email", ""),
                                    "user_name": user.get("name", ""),
                                    "action": "login",
                                    "description": "User logged in",
                                    "record_affected": user.get("username") or user.get("email", ""),
                                    "ip_address": (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
                                                  or (request.client.host if request.client else "n/a"),
                                    "created_at": iso(now())})
    return doc_to_json(user)

@api.post("/auth/register")
async def register(body: RegisterIn, request: Request, current=Depends(get_current_user)):
    if current["role"] != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can create users")
    email = (body.email or "").lower().strip()
    username = body.username.lower().strip()
    if not username:
        raise HTTPException(status_code=400, detail="Username is required")
    if await db.users.find_one({"username": username}):
        raise HTTPException(status_code=400, detail="Username already taken")
    if email and await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    doc = {"username": username, "email": email, "password_hash": hash_password(body.password),
           "name": body.name, "role": body.role, "is_active": True, "created_at": iso(now())}
    res = await db.users.insert_one(doc)
    await audit(current, "user_created", f"Created user {username} ({body.role})", request, username)
    doc["id"] = str(res.inserted_id); doc.pop("password_hash")
    return doc_to_json(doc)

@api.post("/auth/logout")
async def logout(response: Response, user=Depends(get_current_user)):
    clear_auth_cookies(response)
    return {"ok": True}

@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

@api.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_access_token(str(user["_id"]), user["email"], user["role"])
        response.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax",
                            max_age=ACCESS_TTL_MIN * 60, path="/")
        return {"ok": True}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

@api.post("/auth/change-password")
async def change_password(body: ChangePasswordIn, request: Request, user=Depends(get_current_user)):
    record = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not verify_password(body.old_password, record["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one({"_id": ObjectId(user["id"])},
                              {"$set": {"password_hash": hash_password(body.new_password)}})
    await audit(user, "password_change", "User changed password", request, user["email"])
    return {"ok": True}

@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordIn):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if user:
        token = secrets.token_urlsafe(32)
        await db.password_reset_tokens.insert_one({
            "token": token, "user_id": str(user["_id"]),
            "expires_at": iso(now() + timedelta(hours=1)),
            "used": False, "created_at": iso(now())
        })
        log.info(f"Password reset for {email}: token={token}")
    return {"ok": True, "message": "If account exists, a reset token has been generated."}

@api.post("/auth/reset-password")
async def reset_password(body: ResetPasswordIn):
    rec = await db.password_reset_tokens.find_one({"token": body.token, "used": False})
    if not rec:
        raise HTTPException(status_code=400, detail="Invalid or used token")
    exp = rec["expires_at"]
    if isinstance(exp, str):
        exp = datetime.fromisoformat(exp)
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp < now():
        raise HTTPException(status_code=400, detail="Token expired")
    await db.users.update_one({"_id": ObjectId(rec["user_id"])},
                              {"$set": {"password_hash": hash_password(body.new_password)}})
    await db.password_reset_tokens.update_one({"_id": rec["_id"]}, {"$set": {"used": True}})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Users (Admin only)
# ---------------------------------------------------------------------------
@api.get("/users")
async def list_users(user=Depends(require_role(ROLE_ADMIN))):
    rows = await db.users.find({}, {"password_hash": 0}).sort("created_at", -1).to_list(500)
    return [doc_to_json(d) for d in rows]

@api.patch("/users/{user_id}")
async def update_user(user_id: str, payload: dict, request: Request, user=Depends(require_role(ROLE_ADMIN))):
    allowed = {k: v for k, v in payload.items() if k in ("name", "role", "is_active")}
    if payload.get("password"):
        allowed["password_hash"] = hash_password(payload["password"])
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": allowed})
    await audit(user, "user_updated", f"Updated user {user_id}", request, user_id)
    doc = await db.users.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    return doc_to_json(doc)


# ---------------------------------------------------------------------------
# Employees (Admin & Supervisor)
# ---------------------------------------------------------------------------
@api.get("/employees")
async def list_employees(active_only: bool = False, user=Depends(get_current_user)):
    q = {"is_active": True} if active_only else {}
    rows = await db.employees.find(q).sort("name", 1).to_list(1000)
    return [doc_to_json(d) for d in rows]

@api.post("/employees")
async def create_employee(body: EmployeeIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    doc = body.model_dump()
    doc["created_at"] = iso(now()); doc["updated_at"] = iso(now())
    res = await db.employees.insert_one(doc)
    await audit(user, "employee_created", f"Employee {body.name}", request, body.employee_code or body.name)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc

@api.patch("/employees/{eid}")
async def update_employee(eid: str, body: EmployeeIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    data = body.model_dump(); data["updated_at"] = iso(now())
    await db.employees.update_one({"_id": ObjectId(eid)}, {"$set": data})
    await audit(user, "employee_updated", f"Employee {eid}", request, eid)
    return doc_to_json(await db.employees.find_one({"_id": ObjectId(eid)}))

@api.delete("/employees/{eid}")
async def delete_employee(eid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.employees.update_one({"_id": ObjectId(eid)}, {"$set": {"is_active": False, "updated_at": iso(now())}})
    await audit(user, "employee_deactivated", f"Employee {eid}", request, eid)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Categories
# ---------------------------------------------------------------------------
@api.get("/categories")
async def list_categories(include_archived: bool = False, user=Depends(get_current_user)):
    q = {} if include_archived else {"is_archived": {"$ne": True}}
    rows = await db.categories.find(q).sort("name", 1).to_list(500)
    return [doc_to_json(d) for d in rows]

@api.post("/categories")
async def create_category(body: CategoryIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    doc = {"name": body.name.strip(), "description": body.description or "",
           "is_archived": False, "created_at": iso(now()), "updated_at": iso(now())}
    res = await db.categories.insert_one(doc)
    await audit(user, "category_created", f"Category {body.name}", request, body.name)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc

@api.patch("/categories/{cid}")
async def update_category(cid: str, body: CategoryIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.categories.update_one({"_id": ObjectId(cid)},
                                   {"$set": {"name": body.name, "description": body.description,
                                             "updated_at": iso(now())}})
    await audit(user, "category_updated", f"Category {cid}", request, cid)
    return doc_to_json(await db.categories.find_one({"_id": ObjectId(cid)}))

@api.post("/categories/{cid}/archive")
async def archive_category(cid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.categories.update_one({"_id": ObjectId(cid)}, {"$set": {"is_archived": True, "updated_at": iso(now())}})
    await audit(user, "category_archived", f"Category {cid}", request, cid)
    return {"ok": True}

@api.post("/categories/{cid}/restore")
async def restore_category(cid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.categories.update_one({"_id": ObjectId(cid)}, {"$set": {"is_archived": False, "updated_at": iso(now())}})
    await audit(user, "category_restored", f"Category {cid}", request, cid)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Suppliers
# ---------------------------------------------------------------------------
@api.get("/suppliers")
async def list_suppliers(include_archived: bool = False, user=Depends(get_current_user)):
    q = {} if include_archived else {"is_archived": {"$ne": True}}
    rows = await db.suppliers.find(q).sort("name", 1).to_list(500)
    return [doc_to_json(d) for d in rows]

@api.post("/suppliers")
async def create_supplier(body: SupplierIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    doc = body.model_dump(); doc["is_archived"] = False
    doc["created_at"] = iso(now()); doc["updated_at"] = iso(now())
    res = await db.suppliers.insert_one(doc)
    await audit(user, "supplier_created", f"Supplier {body.name}", request, body.name)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc

@api.patch("/suppliers/{sid}")
async def update_supplier(sid: str, body: SupplierIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    data = body.model_dump(); data["updated_at"] = iso(now())
    await db.suppliers.update_one({"_id": ObjectId(sid)}, {"$set": data})
    await audit(user, "supplier_updated", f"Supplier {sid}", request, sid)
    return doc_to_json(await db.suppliers.find_one({"_id": ObjectId(sid)}))

@api.post("/suppliers/{sid}/archive")
async def archive_supplier(sid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.suppliers.update_one({"_id": ObjectId(sid)}, {"$set": {"is_archived": True, "updated_at": iso(now())}})
    await audit(user, "supplier_archived", f"Supplier {sid}", request, sid)
    return {"ok": True}

@api.post("/suppliers/{sid}/restore")
async def restore_supplier(sid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.suppliers.update_one({"_id": ObjectId(sid)}, {"$set": {"is_archived": False, "updated_at": iso(now())}})
    await audit(user, "supplier_restored", f"Supplier {sid}", request, sid)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Catalogs
# ---------------------------------------------------------------------------
def gen_barcode_svg(value: str) -> str:
    cls = pybarcode.get_barcode_class('code128')
    obj = cls(value, writer=SVGWriter())
    buf = io.BytesIO()
    obj.write(buf, options={"module_height": 12.0, "font_size": 8, "text_distance": 3, "quiet_zone": 2})
    return buf.getvalue().decode("utf-8")

def gen_qr_svg(payload: str) -> str:
    factory = qrcode.image.svg.SvgImage
    img = qrcode.make(payload, image_factory=factory, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf)
    return buf.getvalue().decode("utf-8")


@api.get("/catalogs")
async def list_catalogs(
    q: Optional[str] = None,
    status: Optional[str] = None,
    category_id: Optional[str] = None,
    supplier_id: Optional[str] = None,
    fabric_type: Optional[str] = None,
    include_archived: bool = False,
    skip: int = 0,
    limit: int = 50,
    user=Depends(get_current_user)
):
    filt = {}
    if not include_archived:
        filt["is_archived"] = {"$ne": True}
    if status:
        filt["status"] = status
    if category_id:
        filt["category_id"] = category_id
    if supplier_id:
        filt["supplier_id"] = supplier_id
    if fabric_type:
        filt["fabric_type"] = fabric_type
    if q:
        filt["$or"] = [
            {"catalog_code": {"$regex": q, "$options": "i"}},
            {"catalog_name": {"$regex": q, "$options": "i"}},
            {"cat_no": {"$regex": q, "$options": "i"}},
            {"color": {"$regex": q, "$options": "i"}},
            {"qr_value": {"$regex": q, "$options": "i"}},
        ]
    total = await db.catalogs.count_documents(filt)
    rows = await db.catalogs.find(filt).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    items = [strip_buying_price(doc_to_json(d), user["role"]) for d in rows]
    return {"total": total, "items": items}

@api.get("/catalogs/{cid}")
async def get_catalog(cid: str, user=Depends(get_current_user)):
    doc = await db.catalogs.find_one({"_id": ObjectId(cid)})
    if not doc:
        raise HTTPException(404, "Not found")
    return strip_buying_price(doc_to_json(doc), user["role"])

@api.post("/catalogs")
async def create_catalog(body: CatalogIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    if await db.catalogs.find_one({"catalog_code": body.catalog_code}):
        raise HTTPException(400, "Catalog code already exists")
    doc = body.model_dump()
    # only admin can set buying_price
    if user["role"] != ROLE_ADMIN:
        doc.pop("buying_price", None)
    doc.update({
        "barcode_value": body.catalog_code,
        "qr_value": (body.qr_value or "").strip(),  # manual; NOT auto-generated
        "status": "Available",
        "is_archived": False,
        "created_at": iso(now()),
        "updated_at": iso(now()),
        "created_by": user["email"]
    })
    res = await db.catalogs.insert_one(doc)
    await audit(user, "catalog_created", f"Catalog {body.catalog_code}", request, body.catalog_code)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return strip_buying_price(doc, user["role"])

@api.patch("/catalogs/{cid}")
async def update_catalog(cid: str, body: CatalogIn, request: Request, user=Depends(require_role(ROLE_ADMIN))):
    # ONLY ADMIN can edit catalogs
    data = body.model_dump()
    data["updated_at"] = iso(now())
    data["barcode_value"] = body.catalog_code
    data["qr_value"] = (body.qr_value or "").strip()
    existing = await db.catalogs.find_one({"_id": ObjectId(cid)})
    if not existing:
        raise HTTPException(404, "Not found")
    # log price changes specifically
    if existing.get("buying_price") != data.get("buying_price") or existing.get("selling_price") != data.get("selling_price"):
        await audit(user, "price_updated",
                    f"Catalog {existing.get('catalog_code')} buying={data.get('buying_price')} selling={data.get('selling_price')}",
                    request, existing.get("catalog_code", cid))
    await db.catalogs.update_one({"_id": ObjectId(cid)}, {"$set": data})
    await audit(user, "catalog_updated", f"Catalog {cid}", request, existing.get("catalog_code", cid))
    return strip_buying_price(doc_to_json(await db.catalogs.find_one({"_id": ObjectId(cid)})), user["role"])

@api.post("/catalogs/{cid}/archive")
async def archive_catalog(cid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.catalogs.update_one({"_id": ObjectId(cid)},
                                 {"$set": {"is_archived": True, "status": "Archived", "updated_at": iso(now())}})
    await audit(user, "catalog_archived", f"Catalog {cid}", request, cid)
    return {"ok": True}

@api.post("/catalogs/{cid}/restore")
async def restore_catalog(cid: str, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    await db.catalogs.update_one({"_id": ObjectId(cid)},
                                 {"$set": {"is_archived": False, "status": "Available", "updated_at": iso(now())}})
    await audit(user, "catalog_restored", f"Catalog {cid}", request, cid)
    return {"ok": True}

@api.get("/catalogs/{cid}/barcode.svg")
async def catalog_barcode_svg(cid: str, user=Depends(get_current_user)):
    doc = await db.catalogs.find_one({"_id": ObjectId(cid)})
    if not doc:
        raise HTTPException(404, "Not found")
    return Response(content=gen_barcode_svg(doc["barcode_value"]), media_type="image/svg+xml")

@api.get("/catalogs/{cid}/qr.svg")
async def catalog_qr_svg(cid: str, user=Depends(get_current_user)):
    doc = await db.catalogs.find_one({"_id": ObjectId(cid)})
    if not doc:
        raise HTTPException(404, "Not found")
    value = (doc.get("qr_value") or "").strip()
    if not value:
        raise HTTPException(404, "No QR value for this catalog")
    return Response(content=gen_qr_svg(value), media_type="image/svg+xml")

@api.get("/catalogs/{cid}/history")
async def catalog_history(cid: str, user=Depends(get_current_user)):
    issues = await db.catalog_issues.find({"catalog_id": cid}).sort("created_at", -1).to_list(200)
    returns = await db.catalog_returns.find({"catalog_id": cid}).sort("created_at", -1).to_list(200)
    scans = await db.scan_history.find({"catalog_id": cid}).sort("created_at", -1).to_list(200)
    return {
        "issues": [doc_to_json(d) for d in issues],
        "returns": [doc_to_json(d) for d in returns],
        "scans": [doc_to_json(d) for d in scans],
    }


# ---------------------------------------------------------------------------
# Catalog Bulk Import
# ---------------------------------------------------------------------------
IMPORT_COLUMNS = ["catalog_code", "catalog_name", "category", "supplier", "fabric_type",
                  "material_composition", "gsm", "color", "total_swatches", "description",
                  "qr_value", "buying_price", "selling_price"]

@api.get("/catalogs/import/template.xlsx")
async def catalog_import_template(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    wb = Workbook(); ws = wb.active; ws.title = "Catalogs"
    ws.append(IMPORT_COLUMNS)
    ws.append(["FC-001", "Premium Cotton", "Cotton", "Acme Textiles", "Cotton",
               "100% Cotton", 180, "Indigo", 12, "Soft hand feel",
               "QR-FC-001-XYZ", 45.50, 75.00])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=catalog_import_template.xlsx"})

@api.post("/catalogs/import")
async def catalog_import(request: Request,
                         file: UploadFile = File(...),
                         user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    raw = await file.read()
    rows_in = []
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows_in.append(dict(zip(headers, row)))
    except Exception as e:
        raise HTTPException(400, f"Could not parse Excel: {e}")

    # cache categories/suppliers by name for quick lookup
    cats = {c["name"].lower(): str(c["_id"]) async for c in db.categories.find({})}
    sups = {s["name"].lower(): str(s["_id"]) async for s in db.suppliers.find({})}

    success, errors = [], []
    for i, r in enumerate(rows_in, start=2):
        code = (r.get("catalog_code") or "").strip()
        name = (r.get("catalog_name") or "").strip()
        if not code or not name:
            errors.append({"row": i, "error": "catalog_code and catalog_name are required"})
            continue
        if await db.catalogs.find_one({"catalog_code": code}):
            errors.append({"row": i, "catalog_code": code, "error": "Already exists"})
            continue
        cat_name = (r.get("category") or "").strip().lower()
        sup_name = (r.get("supplier") or "").strip().lower()
        cat_id, sup_id = None, None
        if cat_name:
            cat_id = cats.get(cat_name)
            if not cat_id:
                # create category on the fly
                res = await db.categories.insert_one({"name": (r.get("category") or "").strip(),
                                                      "is_archived": False,
                                                      "created_at": iso(now()), "updated_at": iso(now())})
                cat_id = str(res.inserted_id); cats[cat_name] = cat_id
        if sup_name:
            sup_id = sups.get(sup_name)
            if not sup_id:
                res = await db.suppliers.insert_one({"name": (r.get("supplier") or "").strip(),
                                                     "is_archived": False,
                                                     "created_at": iso(now()), "updated_at": iso(now())})
                sup_id = str(res.inserted_id); sups[sup_name] = sup_id

        try:
            gsm = float(r.get("gsm")) if r.get("gsm") not in (None, "") else None
            total_swatches = int(r.get("total_swatches")) if r.get("total_swatches") not in (None, "") else 0
            buying_price = float(r.get("buying_price")) if r.get("buying_price") not in (None, "") else None
            selling_price = float(r.get("selling_price")) if r.get("selling_price") not in (None, "") else None
        except (ValueError, TypeError) as e:
            errors.append({"row": i, "catalog_code": code, "error": f"Numeric field invalid: {e}"})
            continue

        doc = {
            "catalog_code": code, "catalog_name": name,
            "category_id": cat_id, "supplier_id": sup_id,
            "fabric_type": (r.get("fabric_type") or "").strip(),
            "material_composition": (r.get("material_composition") or "").strip(),
            "gsm": gsm, "color": (r.get("color") or "").strip(),
            "total_swatches": total_swatches,
            "description": (r.get("description") or "").strip(),
            "qr_value": (r.get("qr_value") or "").strip(),
            "buying_price": buying_price if user["role"] == ROLE_ADMIN else None,
            "selling_price": selling_price,
            "catalog_image": "", "swatch_images": [],
            "barcode_value": code,
            "status": "Available", "is_archived": False,
            "created_at": iso(now()), "updated_at": iso(now()),
            "created_by": user["email"], "imported": True,
        }
        await db.catalogs.insert_one(doc)
        success.append({"row": i, "catalog_code": code})

    # tracking record
    await db.import_logs.insert_one({
        "user_id": user["id"], "user_email": user["email"],
        "filename": file.filename, "total_rows": len(rows_in),
        "success_count": len(success), "error_count": len(errors),
        "errors": errors[:50], "created_at": iso(now())
    })
    await audit(user, "catalog_import",
                f"Imported {len(success)}/{len(rows_in)} from {file.filename}",
                request, file.filename)
    return {"success": success, "errors": errors,
            "total": len(rows_in), "imported": len(success), "failed": len(errors)}


# ---------------------------------------------------------------------------
# Issues & Returns
# ---------------------------------------------------------------------------
@api.get("/issues")
async def list_issues(filter: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    today = now().date()
    week_end = today + timedelta(days=7)
    if filter == "active":
        q["status"] = "Active"
    elif filter == "due_today":
        q["status"] = "Active"
        q["expected_return_date"] = {"$regex": f"^{today.isoformat()}"}
    elif filter == "due_week":
        q["status"] = "Active"
        q["expected_return_date"] = {"$gte": today.isoformat(), "$lte": week_end.isoformat()}
    elif filter == "overdue":
        q["status"] = "Active"
        q["expected_return_date"] = {"$lt": today.isoformat(), "$ne": ""}
    rows = await db.catalog_issues.find(q).sort("created_at", -1).to_list(2000)
    items = [doc_to_json(d) for d in rows]
    for it in items:
        if it.get("catalog_id"):
            cat = await db.catalogs.find_one({"_id": ObjectId(it["catalog_id"])})
            if cat:
                it["catalog_code"] = cat.get("catalog_code")
                it["catalog_name"] = cat.get("catalog_name")
                it["cat_no"] = cat.get("cat_no", "")
                if cat.get("supplier_id"):
                    sup = await db.suppliers.find_one({"_id": ObjectId(cat["supplier_id"])})
                    it["supplier_name"] = sup.get("name") if sup else ""
        # compute overdue_days
        if it.get("status") == "Active" and it.get("expected_return_date"):
            try:
                due = datetime.fromisoformat(it["expected_return_date"].split("T")[0]).date()
                it["overdue_days"] = max(0, (today - due).days)
                it["is_overdue"] = it["overdue_days"] > 0
            except Exception:
                it["overdue_days"] = 0
                it["is_overdue"] = False
    return items

@api.post("/issues")
async def create_issue(body: IssueIn, request: Request, user=Depends(get_current_user)):
    # all roles can issue
    if not body.mobile or not MOBILE_RE.match(body.mobile.strip()):
        raise HTTPException(status_code=400, detail="Valid mobile number is required (7-20 digits, +-/spaces allowed)")
    cat = await db.catalogs.find_one({"_id": ObjectId(body.catalog_id)})
    if not cat:
        raise HTTPException(404, "Catalog not found")
    if cat.get("status") == "Issued":
        raise HTTPException(400, "Catalog is already issued")
    if cat.get("is_archived"):
        raise HTTPException(400, "Catalog is archived")
    # resolve employee name from id if provided
    employee_name = body.employee_name or ""
    if body.employee_id:
        emp = await db.employees.find_one({"_id": ObjectId(body.employee_id)})
        if emp:
            employee_name = emp.get("name", employee_name)
    doc = body.model_dump()
    doc["transaction_id"] = await next_transaction_id()
    doc["employee_name"] = employee_name
    doc["status"] = "Active"
    doc["issued_by"] = user.get("name") or user.get("username") or user["email"]
    doc["issued_by_email"] = user["email"]
    doc["issue_date"] = doc.get("issue_date") or iso(now())
    doc["created_at"] = iso(now())
    res = await db.catalog_issues.insert_one(doc)
    await db.catalogs.update_one({"_id": ObjectId(body.catalog_id)},
                                 {"$set": {"status": "Issued", "updated_at": iso(now())}})
    await audit(user, "catalog_issued",
                f"{doc['transaction_id']} · {cat['catalog_code']} issued to {body.customer_name or employee_name} ({body.mobile})",
                request, cat['catalog_code'])
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc


@api.get("/returns")
async def list_returns(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await db.catalog_returns.find({}).sort("created_at", -1).to_list(1000)
    items = [doc_to_json(d) for d in rows]
    for it in items:
        if it.get("catalog_id"):
            cat = await db.catalogs.find_one({"_id": ObjectId(it["catalog_id"])})
            if cat:
                it["catalog_code"] = cat.get("catalog_code")
                it["catalog_name"] = cat.get("catalog_name")
        # also bring last issue for mobile/customer
        last_issue = await db.catalog_issues.find_one({"catalog_id": it.get("catalog_id"), "status": {"$in": ["Returned", "Active"]}},
                                                     sort=[("created_at", -1)])
        if last_issue:
            it["customer_name"] = last_issue.get("customer_name")
            it["mobile"] = last_issue.get("mobile")
    return items

@api.post("/returns")
async def create_return(body: ReturnIn, request: Request, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    cat = await db.catalogs.find_one({"_id": ObjectId(body.catalog_id)})
    if not cat:
        raise HTTPException(404, "Catalog not found")
    if cat.get("status") != "Issued":
        raise HTTPException(400, "Catalog is not currently issued")
    doc = body.model_dump()
    doc["transaction_id"] = await next_transaction_id()
    doc["return_date"] = doc.get("return_date") or iso(now())
    doc["received_by"] = user.get("name") or user.get("username") or user["email"]
    doc["received_by_email"] = user["email"]
    doc["created_at"] = iso(now())
    res = await db.catalog_returns.insert_one(doc)
    await db.catalog_issues.update_many({"catalog_id": body.catalog_id, "status": "Active"},
                                        {"$set": {"status": "Returned", "actual_return_date": doc["return_date"]}})
    # CRITICAL FIX: catalog status returns to 'Available' so it can be re-issued
    await db.catalogs.update_one({"_id": ObjectId(body.catalog_id)},
                                 {"$set": {"status": "Available", "updated_at": iso(now())}})
    await audit(user, "catalog_returned",
                f"{doc['transaction_id']} · {cat['catalog_code']} returned ({body.condition})",
                request, cat['catalog_code'])
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc


# ---------------------------------------------------------------------------
# Scans
# ---------------------------------------------------------------------------
@api.post("/scans")
async def scan_barcode(body: ScanIn, request: Request, user=Depends(get_current_user)):
    code = body.barcode_value.strip()
    # try barcode (catalog_code), then qr_value
    cat = await db.catalogs.find_one({"$or": [
        {"barcode_value": code},
        {"catalog_code": code},
        {"qr_value": code},
    ]})
    if not cat:
        raise HTTPException(404, "No catalog matched this code or QR")
    doc = {
        "catalog_id": str(cat["_id"]),
        "user_id": user["id"], "user_email": user["email"],
        "device_type": body.device_type, "action": body.action,
        "remarks": body.remarks, "created_at": iso(now())
    }
    res = await db.scan_history.insert_one(doc)
    await audit(user, "barcode_scan", f"Scanned {cat['catalog_code']} ({body.action})", request, cat['catalog_code'])
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return {"catalog": strip_buying_price(doc_to_json(cat), user["role"]), "scan": doc}

@api.get("/scans")
async def list_scans(limit: int = 100, user=Depends(get_current_user)):
    rows = await db.scan_history.find({}).sort("created_at", -1).limit(limit).to_list(limit)
    items = [doc_to_json(d) for d in rows]
    for it in items:
        if it.get("catalog_id"):
            cat = await db.catalogs.find_one({"_id": ObjectId(it["catalog_id"])})
            if cat:
                it["catalog_code"] = cat.get("catalog_code")
                it["catalog_name"] = cat.get("catalog_name")
    return items


# ---------------------------------------------------------------------------
# Audit / Import logs
# ---------------------------------------------------------------------------
@api.get("/audit-logs")
async def list_audit_logs(limit: int = 300, user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await db.audit_logs.find({}).sort("created_at", -1).limit(limit).to_list(limit)
    return [doc_to_json(d) for d in rows]

@api.get("/import-logs")
async def list_import_logs(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await db.import_logs.find({}).sort("created_at", -1).limit(100).to_list(100)
    return [doc_to_json(d) for d in rows]


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@api.get("/dashboard/stats")
async def dashboard_stats(user=Depends(get_current_user)):
    today = now().date()
    today_iso = today.isoformat()
    week_end_iso = (today + timedelta(days=7)).isoformat()
    today_start = datetime.combine(today, datetime.min.time(), tzinfo=timezone.utc).isoformat()

    total = await db.catalogs.count_documents({"is_archived": {"$ne": True}})
    available = await db.catalogs.count_documents({"status": "Available", "is_archived": {"$ne": True}})
    issued = await db.catalogs.count_documents({"status": "Issued", "is_archived": {"$ne": True}})
    # 'returned' now means number of completed return transactions (catalogs go back to Available after return)
    returned = await db.catalog_returns.count_documents({})
    archived = await db.catalogs.count_documents({"is_archived": True})
    suppliers = await db.suppliers.count_documents({"is_archived": {"$ne": True}})
    categories = await db.categories.count_documents({"is_archived": {"$ne": True}})
    employees = await db.employees.count_documents({"is_active": True})
    scans_today = await db.scan_history.count_documents({"created_at": {"$gte": today_start}})

    due_today = await db.catalog_issues.count_documents({
        "status": "Active",
        "expected_return_date": {"$regex": f"^{today_iso}"}
    })
    due_week = await db.catalog_issues.count_documents({
        "status": "Active",
        "expected_return_date": {"$gte": today_iso, "$lte": week_end_iso}
    })
    overdue_count = await db.catalog_issues.count_documents({
        "status": "Active",
        "expected_return_date": {"$lt": today_iso, "$ne": ""}
    })

    # overdue list with details
    overdue_cursor = db.catalog_issues.find({
        "status": "Active",
        "expected_return_date": {"$lt": today_iso, "$ne": ""}
    }).sort("expected_return_date", 1).limit(20)
    overdue = []
    for it in await overdue_cursor.to_list(20):
        item = doc_to_json(it)
        cat = await db.catalogs.find_one({"_id": ObjectId(item["catalog_id"])})
        if cat:
            item["catalog_code"] = cat.get("catalog_code")
            item["catalog_name"] = cat.get("catalog_name")
        try:
            d = datetime.fromisoformat(item["expected_return_date"].split("T")[0]).date()
            item["overdue_days"] = (today - d).days
        except Exception:
            item["overdue_days"] = 0
        overdue.append(item)

    recent_added = [strip_buying_price(doc_to_json(d), user["role"]) for d in await db.catalogs.find({"is_archived": {"$ne": True}})
                    .sort("created_at", -1).limit(8).to_list(8)]
    recent_returned = [doc_to_json(d) for d in await db.catalog_returns.find({}).sort("created_at", -1).limit(8).to_list(8)]

    return {
        "totals": {
            "total_catalogs": total, "available": available, "issued": issued,
            "returned": returned, "archived": archived,
            "suppliers": suppliers, "categories": categories, "employees": employees,
            "scans_today": scans_today,
            "due_today": due_today, "due_week": due_week, "overdue": overdue_count,
        },
        "recently_added": recent_added,
        "recently_returned": recent_returned,
        "overdue_catalogs": overdue,
    }

@api.get("/dashboard/charts")
async def dashboard_charts(user=Depends(get_current_user)):
    months = []
    today = now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    for i in range(5, -1, -1):
        month = (today.month - i - 1) % 12 + 1
        year = today.year + ((today.month - i - 1) // 12)
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        iss = await db.catalog_issues.count_documents({"created_at": {"$gte": iso(start), "$lt": iso(end)}})
        ret = await db.catalog_returns.count_documents({"created_at": {"$gte": iso(start), "$lt": iso(end)}})
        months.append({"label": start.strftime("%b %Y"), "issues": iss, "returns": ret})

    cat_pipe = [{"$match": {"is_archived": {"$ne": True}}}, {"$group": {"_id": "$category_id", "count": {"$sum": 1}}}]
    cat_dist = []
    async for r in db.catalogs.aggregate(cat_pipe):
        name = "Uncategorized"
        if r["_id"]:
            c = await db.categories.find_one({"_id": ObjectId(r["_id"])})
            if c: name = c["name"]
        cat_dist.append({"name": name, "value": r["count"]})

    sup_pipe = [{"$match": {"is_archived": {"$ne": True}}}, {"$group": {"_id": "$supplier_id", "count": {"$sum": 1}}}]
    sup_dist = []
    async for r in db.catalogs.aggregate(sup_pipe):
        name = "Unknown"
        if r["_id"]:
            s = await db.suppliers.find_one({"_id": ObjectId(r["_id"])})
            if s: name = s["name"]
        sup_dist.append({"name": name, "value": r["count"]})

    issued_pipe = [{"$group": {"_id": "$catalog_id", "count": {"$sum": 1}}},
                   {"$sort": {"count": -1}}, {"$limit": 8}]
    most_issued = []
    async for r in db.catalog_issues.aggregate(issued_pipe):
        if not r["_id"]: continue
        c = await db.catalogs.find_one({"_id": ObjectId(r["_id"])})
        if c: most_issued.append({"name": c.get("catalog_name", c.get("catalog_code", "")), "value": r["count"]})

    scan_pipe = [{"$group": {"_id": "$catalog_id", "count": {"$sum": 1}}},
                 {"$sort": {"count": -1}}, {"$limit": 8}]
    most_scanned = []
    async for r in db.scan_history.aggregate(scan_pipe):
        if not r["_id"]: continue
        c = await db.catalogs.find_one({"_id": ObjectId(r["_id"])})
        if c: most_scanned.append({"name": c.get("catalog_name", c.get("catalog_code", "")), "value": r["count"]})

    return {"monthly": months, "category_distribution": cat_dist,
            "supplier_distribution": sup_dist,
            "most_issued": most_issued, "most_scanned": most_scanned}


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------
async def _catalog_rows(filter_status=None, include_archived=False):
    q = {}
    if filter_status: q["status"] = filter_status
    if not include_archived: q["is_archived"] = {"$ne": True}
    return await db.catalogs.find(q).sort("created_at", -1).to_list(2000)

@api.get("/reports/catalogs/csv")
async def report_catalogs_csv(status: Optional[str] = None, include_archived: bool = False, user=Depends(get_current_user)):
    rows = await _catalog_rows(status, include_archived)
    buf = io.StringIO(); w = csv.writer(buf)
    hdr = ["Code", "Name", "Status", "Fabric", "GSM", "Color", "Swatches", "Selling Price"]
    if user["role"] == ROLE_ADMIN: hdr.insert(7, "Buying Price")
    hdr.append("Created"); w.writerow(hdr)
    for r in rows:
        row = [r.get("catalog_code"), r.get("catalog_name"), r.get("status"),
               r.get("fabric_type"), r.get("gsm"), r.get("color"), r.get("total_swatches"),
               r.get("selling_price")]
        if user["role"] == ROLE_ADMIN: row.insert(7, r.get("buying_price"))
        row.append(r.get("created_at")); w.writerow(row)
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=catalogs.csv"})

@api.get("/reports/catalogs/xlsx")
async def report_catalogs_xlsx(status: Optional[str] = None, include_archived: bool = False, user=Depends(get_current_user)):
    rows = await _catalog_rows(status, include_archived)
    wb = Workbook(); ws = wb.active; ws.title = "Catalogs"
    hdr = ["Code", "Name", "Status", "Fabric", "GSM", "Color", "Swatches", "Selling Price"]
    if user["role"] == ROLE_ADMIN: hdr.insert(7, "Buying Price")
    hdr.append("Created"); ws.append(hdr)
    for r in rows:
        row = [r.get("catalog_code"), r.get("catalog_name"), r.get("status"),
               r.get("fabric_type"), r.get("gsm"), r.get("color"), r.get("total_swatches"),
               r.get("selling_price")]
        if user["role"] == ROLE_ADMIN: row.insert(7, r.get("buying_price"))
        row.append(r.get("created_at")); ws.append(row)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=catalogs.xlsx"})

@api.get("/reports/catalogs/pdf")
async def report_catalogs_pdf(status: Optional[str] = None, include_archived: bool = False, user=Depends(get_current_user)):
    rows = await _catalog_rows(status, include_archived)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Royal Shades — Catalog Report", styles["Title"]),
             Paragraph(f"Generated: {now().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
             Spacer(1, 12)]
    hdr = ["Code", "Name", "Status", "Fabric", "GSM", "Color", "Swatches", "Sell"]
    if user["role"] == ROLE_ADMIN: hdr.insert(7, "Buy")
    data = [hdr]
    for r in rows:
        row = [r.get("catalog_code", ""), r.get("catalog_name", ""), r.get("status", ""),
               r.get("fabric_type", ""), str(r.get("gsm") or ""),
               r.get("color", ""), str(r.get("total_swatches") or 0),
               str(r.get("selling_price") or "")]
        if user["role"] == ROLE_ADMIN: row.insert(7, str(r.get("buying_price") or ""))
        data.append(row)
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=catalogs.pdf"})


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------
async def _issue_rows():
    today = now().date()
    rows = await db.catalog_issues.find({}).sort("created_at", -1).to_list(5000)
    out = []
    for it in rows:
        item = doc_to_json(it)
        if item.get("catalog_id"):
            cat = await db.catalogs.find_one({"_id": ObjectId(item["catalog_id"])})
            if cat:
                item["catalog_code"] = cat.get("catalog_code")
                item["catalog_name"] = cat.get("catalog_name")
                if cat.get("supplier_id"):
                    sup = await db.suppliers.find_one({"_id": ObjectId(cat["supplier_id"])})
                    item["supplier_name"] = sup.get("name") if sup else ""
                else:
                    item["supplier_name"] = ""
        # overdue
        exp = (item.get("expected_return_date") or "")[:10]
        is_overdue = False
        if item.get("status") == "Active" and exp:
            try:
                due = datetime.fromisoformat(exp).date()
                is_overdue = today > due
            except Exception:
                pass
        item["is_overdue"] = "Yes" if is_overdue else "No"
        item["is_available"] = "No" if item.get("status") == "Active" else "Yes"
        out.append(item)
    return out

@api.get("/reports/issues/csv")
async def report_issues_csv(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await _issue_rows()
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(["Txn ID", "Catalog Code", "Catalog Name", "Supplier", "Customer Name",
                "Employee Name", "Mobile", "Issue Date", "Due Date", "Is Overdue",
                "Is Available", "Issued By", "Status"])
    for r in rows:
        w.writerow([r.get("transaction_id", ""), r.get("catalog_code", ""), r.get("catalog_name", ""),
                    r.get("supplier_name", ""), r.get("customer_name", ""),
                    r.get("employee_name", ""), r.get("mobile", ""),
                    (r.get("issue_date") or "")[:10], (r.get("expected_return_date") or "")[:10],
                    r.get("is_overdue", ""), r.get("is_available", ""),
                    r.get("issued_by", ""), r.get("status", "")])
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=issues_report.csv"})

@api.get("/reports/issues/xlsx")
async def report_issues_xlsx(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await _issue_rows()
    wb = Workbook(); ws = wb.active; ws.title = "Issues"
    ws.append(["Txn ID", "Catalog Code", "Catalog Name", "Supplier", "Customer Name",
               "Employee Name", "Mobile", "Issue Date", "Due Date", "Is Overdue",
               "Is Available", "Issued By", "Status"])
    for r in rows:
        ws.append([r.get("transaction_id", ""), r.get("catalog_code", ""), r.get("catalog_name", ""),
                   r.get("supplier_name", ""), r.get("customer_name", ""),
                   r.get("employee_name", ""), r.get("mobile", ""),
                   (r.get("issue_date") or "")[:10], (r.get("expected_return_date") or "")[:10],
                   r.get("is_overdue", ""), r.get("is_available", ""),
                   r.get("issued_by", ""), r.get("status", "")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=issues_report.xlsx"})

@api.get("/reports/issues/pdf")
async def report_issues_pdf(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    rows = await _issue_rows()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Royal Shades — Issue Report", styles["Title"]),
             Paragraph(f"Generated: {now().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
             Spacer(1, 12)]
    hdr = ["Txn ID", "Code", "Catalog", "Supplier", "Customer", "Employee",
           "Mobile", "Issue", "Due", "Overdue", "Avail", "Status"]
    data = [hdr]
    for r in rows:
        data.append([r.get("transaction_id", ""), r.get("catalog_code", "")[:8],
                     (r.get("catalog_name", "") or "")[:18], (r.get("supplier_name", "") or "")[:12],
                     (r.get("customer_name", "") or "")[:12], (r.get("employee_name", "") or "")[:12],
                     r.get("mobile", ""), (r.get("issue_date") or "")[:10],
                     (r.get("expected_return_date") or "")[:10],
                     r.get("is_overdue", ""), r.get("is_available", ""), r.get("status", "")])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=issues_report.pdf"})

@api.get("/reports/employee-wise")
async def report_employee_wise(user=Depends(require_role(ROLE_ADMIN, ROLE_SUPERVISOR))):
    pipe = [
        {"$group": {"_id": "$employee_name", "total": {"$sum": 1},
                    "active": {"$sum": {"$cond": [{"$eq": ["$status", "Active"]}, 1, 0]}},
                    "returned": {"$sum": {"$cond": [{"$eq": ["$status", "Returned"]}, 1, 0]}}}},
        {"$sort": {"total": -1}}
    ]
    rows = []
    async for r in db.catalog_issues.aggregate(pipe):
        rows.append({"employee_name": r["_id"] or "—",
                     "total_issues": r["total"], "active": r["active"], "returned": r["returned"]})
    return rows


@api.get("/")
async def root():
    return {"message": "Royal Shades Catalog API", "version": "2.1"}


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
@app.on_event("startup")
async def on_startup():
    try:
        # drop legacy unique email index if present (it's now sparse non-unique)
        try:
            await db.users.drop_index("email_1")
        except Exception:
            pass
        await db.users.create_index("username", unique=True, sparse=True)
        await db.users.create_index("email", sparse=True)
        await db.login_attempts.create_index("identifier")
        await db.catalogs.create_index("catalog_code", unique=True)
        await db.catalogs.create_index("barcode_value")
        await db.catalogs.create_index("qr_value")
        await db.catalogs.create_index("cat_no")
        await db.catalogs.create_index("status")
        await db.scan_history.create_index("created_at")
        await db.audit_logs.create_index("created_at")
        await db.employees.create_index("name")
        await db.catalog_issues.create_index("expected_return_date")
        await db.catalog_issues.create_index("status")
        await db.catalog_issues.create_index("transaction_id", unique=True, sparse=True)
        await db.catalog_returns.create_index("transaction_id", unique=True, sparse=True)
    except Exception as e:
        log.warning(f"Index setup: {e}")

    # Migrate role: manager -> supervisor
    res = await db.users.update_many({"role": "manager"}, {"$set": {"role": "supervisor"}})
    if res.modified_count:
        log.info(f"Migrated {res.modified_count} user(s) from manager → supervisor")

    # Backfill username for users that have email but no username
    async for u in db.users.find({"username": {"$exists": False}}):
        email = (u.get("email") or "").lower()
        if not email:
            continue
        base = email.split("@")[0]
        cand = base
        i = 0
        while await db.users.find_one({"username": cand}):
            i += 1
            cand = f"{base}{i}"
        await db.users.update_one({"_id": u["_id"]}, {"$set": {"username": cand}})
        log.info(f"Backfilled username '{cand}' for {email}")

    # Seed admin (using username 'admin' from ADMIN_USERNAME, plus an email for back-compat)
    admin_username = os.environ.get("ADMIN_USERNAME", "admin")
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"$or": [{"username": admin_username}, {"email": admin_email}]})
    if not existing:
        await db.users.insert_one({
            "username": admin_username, "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "System Admin", "role": ROLE_ADMIN, "is_active": True,
            "created_at": iso(now())
        })
        log.info(f"Seeded admin: {admin_username}")
    else:
        updates = {}
        if not existing.get("username"):
            updates["username"] = admin_username
        if not verify_password(admin_password, existing["password_hash"]):
            updates["password_hash"] = hash_password(admin_password)
        if updates:
            await db.users.update_one({"_id": existing["_id"]}, {"$set": updates})

@app.on_event("shutdown")
async def on_shutdown():
    client.close()


app.include_router(api)

frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
allow = [frontend_url]
extra = os.environ.get("CORS_ORIGINS", "")
if extra and extra != "*":
    allow += [o.strip() for o in extra.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=allow,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
