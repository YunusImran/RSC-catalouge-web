"""
Royal Shades Catalog v2.2 Backend Regression Suite

Coverage:
- Re-import: 1315 catalogs with RSC-NNNNN codes; catalog_code == barcode_value == qr_value.
- Catalog list pagination (skip/limit), search by cat_no & name.
- QR svg endpoint serves QR (qr_value populated).
- POST /api/scans accepts RSC-NNNNN value.
- NEW: POST /api/issues/batch — one shared transaction_id; atomicity when a cat is already issued.
- /api/issues batch row carries transaction_id; return one item -> Available, others stay Issued.
- Reports CSV/XLSX/PDF use "Cat No" header (NOT "Catalog Code").
- Dashboard totals sum to ~1315.
- Single POST /api/issues still works and is independent from batch txn.
"""
import os
import io
import re
import time
import pytest
import requests

BASE = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "Admin@123"

TXN_RE = re.compile(r"^TXN-\d{4,}$")


def _clear_lockouts():
    try:
        from pymongo import MongoClient
        m = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
        m[os.environ.get("DB_NAME", "fabric_catalog_db")].login_attempts.delete_many({})
    except Exception:
        pass


def _login(username, password):
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE}/api/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, f"login {username}: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="session")
def admin():
    _clear_lockouts()
    return _login(ADMIN_USERNAME, ADMIN_PASSWORD)


@pytest.fixture(scope="session")
def available_ids(admin):
    """Three available catalog ids for batch tests."""
    r = admin.get(f"{BASE}/api/catalogs?status=Available&limit=5")
    items = r.json()["items"]
    assert len(items) >= 3, f"need >=3 available catalogs, got {len(items)}"
    return [it["id"] for it in items[:3]], [it["catalog_code"] for it in items[:3]]


# ====================== Catalog import / pagination ======================
class TestCatalogImport:
    def test_total_catalogs_1315(self, admin):
        r = admin.get(f"{BASE}/api/catalogs?limit=2")
        assert r.status_code == 200
        body = r.json()
        assert body.get("total") == 1315, f"expected total=1315, got {body.get('total')}"
        assert isinstance(body.get("items"), list)
        assert len(body["items"]) == 2

    def test_first_code_rsc(self, admin):
        r = admin.get(f"{BASE}/api/catalogs?limit=5")
        items = r.json()["items"]
        assert any(it["catalog_code"].startswith("RSC-") for it in items), \
            [it["catalog_code"] for it in items]

    def test_pagination_skip(self, admin):
        a = admin.get(f"{BASE}/api/catalogs?skip=0&limit=10").json()["items"]
        b = admin.get(f"{BASE}/api/catalogs?skip=100&limit=10").json()["items"]
        a_codes = {it["catalog_code"] for it in a}
        b_codes = {it["catalog_code"] for it in b}
        assert a_codes and b_codes
        assert a_codes != b_codes, "skip=0 and skip=100 returned same catalogs"

    def test_search_by_cat_no(self, admin):
        r = admin.get(f"{BASE}/api/catalogs?q=BX-440A&limit=10")
        assert r.status_code == 200
        items = r.json()["items"]
        assert items, "expected a catalog with cat_no BX-440A"
        assert any((it.get("cat_no") or "").upper() == "BX-440A" for it in items)

    def test_search_by_name_hanger(self, admin):
        r = admin.get(f"{BASE}/api/catalogs?q=HANGER&limit=20")
        items = r.json()["items"]
        assert items, "expected HANGER catalogs"
        assert any("HANGER" in (it.get("catalog_name") or "").upper() for it in items)

    def test_code_equals_barcode_equals_qr(self, admin):
        # Pick a few catalogs and confirm catalog_code == barcode_value == qr_value
        r = admin.get(f"{BASE}/api/catalogs?limit=10")
        items = r.json()["items"]
        checked = 0
        for it in items:
            cid = it["id"]
            # GET single catalog (list might trim fields)
            d = admin.get(f"{BASE}/api/catalogs/{cid}").json()
            cc = d.get("catalog_code")
            bv = d.get("barcode_value")
            qv = d.get("qr_value")
            assert cc and bv and qv, f"missing one of code/barcode/qr: {cc}/{bv}/{qv}"
            assert cc == bv == qv, f"mismatch: code={cc} barcode={bv} qr={qv}"
            checked += 1
            if checked >= 3:
                break
        assert checked >= 3

    def test_qr_svg(self, admin):
        r = admin.get(f"{BASE}/api/catalogs?limit=1")
        cid = r.json()["items"][0]["id"]
        rr = admin.get(f"{BASE}/api/catalogs/{cid}/qr.svg")
        assert rr.status_code == 200, rr.text
        assert "svg" in rr.headers.get("content-type", "").lower()
        assert b"<svg" in rr.content[:400]


# ====================== Scans ======================
class TestScanRSC:
    def test_scan_by_rsc_code(self, admin):
        r = admin.post(f"{BASE}/api/scans",
                       json={"barcode_value": "RSC-00002", "action": "Search"})
        assert r.status_code == 200, r.text
        cat = r.json()["catalog"]
        assert cat["catalog_code"] == "RSC-00002"


# ====================== BATCH ISSUE ======================
class TestBatchIssue:
    def test_batch_issue_shared_txn(self, admin, available_ids):
        ids, codes = available_ids
        body = {
            "catalog_ids": ids,
            "customer_name": "Batch Customer",
            "mobile": "+971559908586",
            "expected_return_date": "2026-12-31",
        }
        r = admin.post(f"{BASE}/api/issues/batch", json=body)
        assert r.status_code == 200, r.text
        data = r.json()
        assert TXN_RE.match(data["transaction_id"]), data["transaction_id"]
        assert data["count"] == 3
        assert len(data["issued"]) == 3
        # all 3 share the SAME transaction_id
        txn_ids = {i["transaction_id"] for i in data["issued"]}
        assert txn_ids == {data["transaction_id"]}
        # catalog statuses should be Issued
        for cid in ids:
            c = admin.get(f"{BASE}/api/catalogs/{cid}").json()
            assert c["status"] == "Issued", c
        # store for next tests via pytest cache
        pytest._batch_ids = ids
        pytest._batch_codes = codes
        pytest._batch_txn = data["transaction_id"]

    def test_issues_list_carries_txn(self, admin):
        r = admin.get(f"{BASE}/api/issues?limit=50")
        items = r.json()
        # at least 3 items with our txn id
        same = [it for it in items if it.get("transaction_id") == pytest._batch_txn]
        assert len(same) == 3, f"expected 3 batched items, got {len(same)}"

    def test_batch_with_already_issued_returns_400_atomic(self, admin):
        # find one available and combine with an already-issued one
        avail = admin.get(f"{BASE}/api/catalogs?status=Available&limit=2").json()["items"]
        assert len(avail) >= 2
        body = {
            "catalog_ids": [avail[0]["id"], pytest._batch_ids[0]],  # second one is already Issued
            "customer_name": "Atomic Test",
            "mobile": "+971501234567",
            "expected_return_date": "2026-12-31",
        }
        r = admin.post(f"{BASE}/api/issues/batch", json=body)
        assert r.status_code == 400, r.text
        # error should mention the already-issued catalog code
        detail = r.json().get("detail", "")
        assert pytest._batch_codes[0] in detail, f"expected code {pytest._batch_codes[0]} in: {detail}"
        # atomicity: avail[0] must still be Available
        c = admin.get(f"{BASE}/api/catalogs/{avail[0]['id']}").json()
        assert c["status"] == "Available", f"non-atomic: {avail[0]['catalog_code']} got {c['status']}"

    def test_batch_invalid_mobile_400(self, admin):
        avail = admin.get(f"{BASE}/api/catalogs?status=Available&limit=1").json()["items"]
        r = admin.post(f"{BASE}/api/issues/batch",
                       json={"catalog_ids": [avail[0]["id"]],
                             "customer_name": "X", "mobile": "abc",
                             "expected_return_date": "2026-12-31"})
        assert r.status_code == 400

    def test_batch_empty_ids_400(self, admin):
        r = admin.post(f"{BASE}/api/issues/batch",
                       json={"catalog_ids": [],
                             "customer_name": "X", "mobile": "+971501234567",
                             "expected_return_date": "2026-12-31"})
        assert r.status_code == 400

    def test_return_one_keeps_others_issued(self, admin):
        # Return the first of the 3 batched items
        first_id = pytest._batch_ids[0]
        r = admin.post(f"{BASE}/api/returns",
                       json={"catalog_id": first_id, "condition": "Good"})
        assert r.status_code == 200, r.text
        # first -> Available
        assert admin.get(f"{BASE}/api/catalogs/{first_id}").json()["status"] == "Available"
        # other two -> still Issued
        for cid in pytest._batch_ids[1:]:
            assert admin.get(f"{BASE}/api/catalogs/{cid}").json()["status"] == "Issued"

    def test_single_issue_independent_txn(self, admin):
        avail = admin.get(f"{BASE}/api/catalogs?status=Available&limit=1").json()["items"]
        r = admin.post(f"{BASE}/api/issues",
                       json={"catalog_id": avail[0]["id"],
                             "customer_name": "Solo",
                             "mobile": "+971501231234",
                             "expected_return_date": "2026-12-31"})
        assert r.status_code == 200, r.text
        d = r.json()
        assert TXN_RE.match(d["transaction_id"])
        assert d["transaction_id"] != pytest._batch_txn


# ====================== Reports: "Cat No" header ======================
class TestReportsCatNoHeader:
    def test_csv_has_cat_no(self, admin):
        r = admin.get(f"{BASE}/api/reports/issues/csv")
        assert r.status_code == 200, r.text
        header = r.text.splitlines()[0]
        assert "Cat No" in header, f"expected 'Cat No' in header: {header}"
        assert "Catalog Code" not in header, f"'Catalog Code' should be replaced by 'Cat No': {header}"

    def test_xlsx_has_cat_no(self, admin):
        from openpyxl import load_workbook
        r = admin.get(f"{BASE}/api/reports/issues/xlsx")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Cat No" in header_row, header_row
        assert "Catalog Code" not in header_row, header_row

    def test_pdf_renders(self, admin):
        r = admin.get(f"{BASE}/api/reports/issues/pdf")
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"


# ====================== Dashboard totals ======================
class TestDashboardTotals:
    def test_total_catalogs_1315(self, admin):
        r = admin.get(f"{BASE}/api/dashboard/stats")
        assert r.status_code == 200
        body = r.json()
        # find total_catalogs in either flat or nested form
        total = (body.get("totals") or {}).get("total_catalogs") \
            or body.get("total_catalogs") \
            or (body.get("totals") or {}).get("catalogs")
        assert total == 1315, f"expected 1315, got {total}; body={body}"
