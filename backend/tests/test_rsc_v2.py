"""
Royal Shades Catalog v2 - Backend Regression Suite
Tests employees CRUD, role rename, bulk import, price gating, mobile validation,
returns RBAC, dashboard widgets, scan-by-qr, audit logs.
"""
import os
import io
import time
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")  # public URL
ADMIN_EMAIL = "admin@fabriccatalog.com"
ADMIN_PASS = "Admin@123"
SUP_EMAIL = "supervisor1@test.com"
SUP_PASS = "Sup@1234"
STAFF_EMAIL = "staff1@test.com"
STAFF_PASS = "Stf@1234"


def _login(email, password):
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE}/api/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, f"login {email}: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="session")
def admin():
    # Ensure no lockout from previous tests
    try:
        from pymongo import MongoClient
        m = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
        m[os.environ.get("DB_NAME", "fabric_catalog_db")].login_attempts.delete_many({})
    except Exception:
        pass
    return _login(ADMIN_EMAIL, ADMIN_PASS)


@pytest.fixture(scope="session")
def supervisor(admin):
    # ensure supervisor user exists
    r = admin.post(f"{BASE}/api/auth/register",
                   json={"email": SUP_EMAIL, "password": SUP_PASS, "name": "Sup One", "role": "supervisor"})
    assert r.status_code in (200, 400), r.text
    return _login(SUP_EMAIL, SUP_PASS)


@pytest.fixture(scope="session")
def staff(admin):
    r = admin.post(f"{BASE}/api/auth/register",
                   json={"email": STAFF_EMAIL, "password": STAFF_PASS, "name": "Staff One", "role": "staff"})
    assert r.status_code in (200, 400), r.text
    return _login(STAFF_EMAIL, STAFF_PASS)


# ---------- Role migration ----------
class TestRoleMigration:
    def test_no_manager_role_remains(self, admin):
        r = admin.get(f"{BASE}/api/users")
        assert r.status_code == 200
        managers = [u for u in r.json() if u.get("role") == "manager"]
        assert managers == [], f"Still found manager users: {managers}"

    def test_register_manager_role_rejected(self, admin):
        r = admin.post(f"{BASE}/api/auth/register",
                       json={"email": "manager2@test.com", "password": "x" * 8, "name": "x", "role": "manager"})
        assert r.status_code == 422, r.text  # pydantic Literal mismatch

    def test_register_supervisor_role_ok(self, admin):
        email = f"sup_{int(time.time())}@test.com"
        r = admin.post(f"{BASE}/api/auth/register",
                       json={"email": email, "password": "Sup@12345", "name": "S", "role": "supervisor"})
        assert r.status_code == 200, r.text
        assert r.json()["role"] == "supervisor"


# ---------- Employees CRUD ----------
class TestEmployees:
    def test_create_list_update_delete(self, admin, supervisor):
        # create as admin
        body = {"name": "TEST_Emp1", "employee_code": "E001", "department": "Sales",
                "designation": "Rep", "mobile": "+971501112233", "email": "e1@test.com"}
        r = admin.post(f"{BASE}/api/employees", json=body)
        assert r.status_code == 200, r.text
        eid = r.json()["id"]

        # supervisor can list
        r = supervisor.get(f"{BASE}/api/employees")
        assert r.status_code == 200
        assert any(e["id"] == eid for e in r.json())

        # update as supervisor
        body["designation"] = "Senior Rep"
        r = supervisor.patch(f"{BASE}/api/employees/{eid}", json=body)
        assert r.status_code == 200
        assert r.json()["designation"] == "Senior Rep"

        # delete (soft) as admin
        r = admin.delete(f"{BASE}/api/employees/{eid}")
        assert r.status_code == 200

        # employee still exists but inactive
        r = admin.get(f"{BASE}/api/employees?active_only=true")
        assert all(e["id"] != eid for e in r.json())

    def test_staff_cannot_create_employee(self, staff):
        r = staff.post(f"{BASE}/api/employees", json={"name": "X"})
        assert r.status_code == 403


# ---------- Catalogs: prices + QR ----------
@pytest.fixture(scope="session")
def sample_catalog(admin):
    code = f"RSC-{int(time.time())}"
    body = {"catalog_code": code, "catalog_name": "RSC Sample",
            "buying_price": 10.0, "selling_price": 20.0, "qr_value": "QR-XYZ-9999"}
    r = admin.post(f"{BASE}/api/catalogs", json=body)
    assert r.status_code == 200, r.text
    return r.json()


class TestCatalogPricesQR:
    def test_admin_sees_buying_price(self, admin, sample_catalog):
        r = admin.get(f"{BASE}/api/catalogs/{sample_catalog['id']}")
        assert r.status_code == 200
        d = r.json()
        assert d.get("buying_price") == 10.0
        assert d.get("selling_price") == 20.0
        assert d.get("qr_value") == "QR-XYZ-9999"

    def test_supervisor_no_buying_price(self, supervisor, sample_catalog):
        r = supervisor.get(f"{BASE}/api/catalogs/{sample_catalog['id']}")
        assert r.status_code == 200
        d = r.json()
        assert "buying_price" not in d, f"supervisor should not see buying_price: {d}"
        assert d.get("selling_price") == 20.0

    def test_staff_no_buying_price_list(self, staff, sample_catalog):
        r = staff.get(f"{BASE}/api/catalogs")
        assert r.status_code == 200
        for item in r.json()["items"]:
            assert "buying_price" not in item

    def test_qr_svg_returns_when_present(self, admin, sample_catalog):
        r = admin.get(f"{BASE}/api/catalogs/{sample_catalog['id']}/qr.svg")
        assert r.status_code == 200
        assert "svg" in r.headers.get("content-type", "")

    def test_qr_svg_404_when_empty(self, admin):
        # create catalog WITHOUT qr_value
        code = f"NOQR-{int(time.time())}"
        r = admin.post(f"{BASE}/api/catalogs",
                       json={"catalog_code": code, "catalog_name": "NoQR", "selling_price": 5})
        assert r.status_code == 200
        cid = r.json()["id"]
        r = admin.get(f"{BASE}/api/catalogs/{cid}/qr.svg")
        assert r.status_code == 404

    def test_supervisor_cannot_update_catalog(self, supervisor, sample_catalog):
        r = supervisor.patch(f"{BASE}/api/catalogs/{sample_catalog['id']}",
                             json={"catalog_code": sample_catalog["catalog_code"],
                                   "catalog_name": "X", "selling_price": 99})
        assert r.status_code == 403

    def test_staff_cannot_update_catalog(self, staff, sample_catalog):
        r = staff.patch(f"{BASE}/api/catalogs/{sample_catalog['id']}",
                        json={"catalog_code": sample_catalog["catalog_code"],
                              "catalog_name": "X", "selling_price": 99})
        assert r.status_code == 403


# ---------- Bulk Import ----------
class TestBulkImport:
    def test_template_download(self, admin):
        r = admin.get(f"{BASE}/api/catalogs/import/template.xlsx")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        headers = [c.value for c in wb.active[1]]
        for must in ("catalog_code", "catalog_name", "qr_value", "buying_price", "selling_price"):
            assert must in headers, f"missing {must} in template"

    def test_import_valid_rows(self, admin):
        wb = Workbook(); ws = wb.active
        ws.append(["catalog_code", "catalog_name", "category", "supplier", "fabric_type",
                   "material_composition", "gsm", "color", "total_swatches", "description",
                   "qr_value", "buying_price", "selling_price"])
        ts = int(time.time())
        ws.append([f"IMP-{ts}-A", "Imp A", "Cotton", "Acme", "Cotton", "100% Cotton",
                   180, "Red", 10, "d", f"QR-IMP-{ts}-A", 12.5, 25.0])
        ws.append([f"IMP-{ts}-B", "Imp B", "Cotton", "Acme", "Cotton", "100% Cotton",
                   190, "Blue", 12, "d", f"QR-IMP-{ts}-B", 13.5, 26.0])
        # missing fields row
        ws.append(["", "MissingCode", "", "", "", "", None, "", None, "", "", None, None])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        files = {"file": ("import.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        # use a copy of session WITHOUT JSON content-type
        sess = requests.Session()
        sess.cookies.update(admin.cookies)
        r = sess.post(f"{BASE}/api/catalogs/import", files=files)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["imported"] == 2, data
        assert data["failed"] >= 1
        assert any("required" in e.get("error", "").lower() for e in data["errors"])

    def test_staff_cannot_import(self, staff):
        sess = requests.Session()
        sess.cookies.update(staff.cookies)
        files = {"file": ("x.xlsx", b"x",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = sess.post(f"{BASE}/api/catalogs/import", files=files)
        assert r.status_code == 403


# ---------- Issues mobile validation + overdue ----------
class TestIssues:
    def _make_catalog(self, admin, qr=""):
        code = f"ISS-{int(time.time()*1000)}"
        r = admin.post(f"{BASE}/api/catalogs",
                       json={"catalog_code": code, "catalog_name": "Iss", "qr_value": qr,
                             "selling_price": 10})
        assert r.status_code == 200, r.text
        return r.json()["id"]

    def test_mobile_required(self, admin):
        cid = self._make_catalog(admin)
        r = admin.post(f"{BASE}/api/issues",
                       json={"catalog_id": cid, "customer_name": "C", "mobile": ""})
        assert r.status_code == 400

    def test_mobile_invalid_chars(self, admin):
        cid = self._make_catalog(admin)
        r = admin.post(f"{BASE}/api/issues",
                       json={"catalog_id": cid, "customer_name": "C", "mobile": "12abc"})
        assert r.status_code == 400

    def test_mobile_valid_creates(self, admin):
        cid = self._make_catalog(admin)
        r = admin.post(f"{BASE}/api/issues",
                       json={"catalog_id": cid, "customer_name": "C",
                             "mobile": "+971559908586",
                             "expected_return_date": "2020-01-01"})
        assert r.status_code == 200, r.text
        assert r.json()["mobile"] == "+971559908586"

    def test_overdue_filter(self, admin):
        # previous test created one with date 2020-01-01 => overdue
        r = admin.get(f"{BASE}/api/issues?filter=overdue")
        assert r.status_code == 200
        items = r.json()
        assert len(items) >= 1
        for it in items:
            assert it.get("is_overdue") is True
            assert it.get("overdue_days", 0) > 0


# ---------- Returns RBAC ----------
class TestReturnsRBAC:
    def test_staff_blocked_post_return(self, staff):
        r = staff.post(f"{BASE}/api/returns",
                       json={"catalog_id": "000000000000000000000000", "condition": "Good"})
        assert r.status_code == 403

    def test_staff_blocked_list_returns(self, staff):
        r = staff.get(f"{BASE}/api/returns")
        assert r.status_code == 403

    def test_supervisor_can_list(self, supervisor):
        r = supervisor.get(f"{BASE}/api/returns")
        assert r.status_code == 200


# ---------- Scan by QR ----------
class TestScanQR:
    def test_scan_by_qr_value(self, admin, sample_catalog):
        r = admin.post(f"{BASE}/api/scans",
                       json={"barcode_value": sample_catalog["qr_value"], "action": "Search"})
        assert r.status_code == 200, r.text
        cat = r.json()["catalog"]
        assert cat["id"] == sample_catalog["id"]

    def test_scan_by_qr_supervisor_no_buying(self, supervisor, sample_catalog):
        r = supervisor.post(f"{BASE}/api/scans",
                            json={"barcode_value": sample_catalog["qr_value"], "action": "Search"})
        assert r.status_code == 200
        assert "buying_price" not in r.json()["catalog"]


# ---------- Dashboard ----------
class TestDashboard:
    def test_stats_contains_widgets(self, admin):
        r = admin.get(f"{BASE}/api/dashboard/stats")
        assert r.status_code == 200
        t = r.json()["totals"]
        for k in ("due_today", "due_week", "overdue", "issued", "returned"):
            assert k in t, f"missing {k}"
        # overdue list shape
        for oc in r.json().get("overdue_catalogs", []):
            assert "overdue_days" in oc


# ---------- Users admin-only ----------
class TestUsersACL:
    def test_supervisor_blocked_list_users(self, supervisor):
        r = supervisor.get(f"{BASE}/api/users")
        assert r.status_code == 403

    def test_admin_can_list_users(self, admin):
        r = admin.get(f"{BASE}/api/users")
        assert r.status_code == 200


# ---------- Audit logs ----------
class TestAudit:
    def test_audit_has_recent_actions(self, admin):
        r = admin.get(f"{BASE}/api/audit-logs?limit=300")
        assert r.status_code == 200
        actions = {e.get("action") for e in r.json()}
        # employee_created and catalog_import should be present due to earlier tests
        assert "employee_created" in actions
        assert "catalog_import" in actions
        # Newly created audit entries must carry user_name and record_affected
        new_entries = [e for e in r.json() if e.get("action") in ("catalog_import", "employee_created")]
        assert new_entries, "no catalog_import/employee_created entries found"
        for e in new_entries[:5]:
            assert "user_name" in e, f"missing user_name: {e}"
            assert "record_affected" in e, f"missing record_affected: {e}"
